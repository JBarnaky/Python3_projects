from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl

fp = webdriver.FirefoxProfile()
fp.set_preference("browser.helperApps.neverAsk.saveToDisk", "text/plain,application/pdf")
fp.set_preference("browser.download.manager.showWhenStarting", False)
fp.set_preference("browser.download.dir", "C:\DownloadedFiles")
fp.set_preference("browser.download.folderList", 2)
fp.set_preference("pdfjs.disabled", True)

driver = webdriver.Firefox(executable_path = "D:\geckodriver-v0.27.0-win64\geckodriver.exe", firefox_profile=fp)

driver.maximize_window()

# driver.get("https://www.youtube.com/")

path = "D:/test1.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(row=r, column=c).value, end="     ")

    print()

for r in range(1, 5):
    for c in range(1, 4):
        sheet.cell(row=r, column=c).value = "welcome"

workbook.save(path)